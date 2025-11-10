#!/usr/bin/env python3
"""
Substack Broken Link Checker
Checks links from Substack posts and generates a CSV report of broken links.
"""

import csv
import re
import time
import sys
from datetime import datetime
from typing import List, Dict, Tuple
from urllib.parse import urljoin, urlparse
import xml.etree.ElementTree as ET

import requests
import cloudscraper
from bs4 import BeautifulSoup

try:
    from playwright.sync_api import sync_playwright
    PLAYWRIGHT_AVAILABLE = True
except ImportError:
    PLAYWRIGHT_AVAILABLE = False
    print("Warning: Playwright not available. Install with: pip install playwright && playwright install")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not available. Excel output disabled. Install with: pip install openpyxl")


class SubstackLinkChecker:
    def __init__(self, base_url: str, timeout: int = 10, use_playwright: bool = True, cookie: str = None):
        self.base_url = base_url
        self.timeout = timeout
        self.use_playwright = use_playwright and PLAYWRIGHT_AVAILABLE
        self.cookie = cookie

        if not self.use_playwright:
            # Fallback to cloudscraper if Playwright not available
            self.session = cloudscraper.create_scraper(
                browser={
                    'browser': 'chrome',
                    'platform': 'darwin',
                    'desktop': True
                }
            )
            # Add additional headers to make requests more realistic
            headers = {
                'Accept-Language': 'en-US,en;q=0.9',
                'Cache-Control': 'max-age=0',
                'Referer': 'https://www.google.com/',
            }

            # If cookie provided (substack.sid or connect.sid from logged-in session), add it to headers
            # This bypasses bot protection by using authenticated session
            if self.cookie:
                # Support both cookie names (Substack uses substack.sid, some may use connect.sid)
                headers['Cookie'] = f'substack.sid={self.cookie}'
                print(f"Using authenticated session cookie for requests")

            self.session.headers.update(headers)
        else:
            self.session = cloudscraper.create_scraper()  # Still used for link checking

        self.results = []

    def fetch_sitemap(self) -> List[str]:
        """Fetch and parse the sitemap to get post URLs."""
        sitemap_url = f"{self.base_url}/sitemap.xml"
        print(f"Fetching sitemap from {sitemap_url}...")

        try:
            response = self.session.get(sitemap_url, timeout=self.timeout)
            response.raise_for_status()

            # Parse XML
            root = ET.fromstring(response.content)

            # Handle sitemap index (if it exists)
            namespace = {'ns': 'http://www.sitemaps.org/schemas/sitemap/0.9'}

            # Check if this is a sitemap index
            sitemaps = root.findall('.//ns:sitemap/ns:loc', namespace)
            if sitemaps:
                print(f"Found sitemap index with {len(sitemaps)} sitemaps")
                return [sitemap.text for sitemap in sitemaps]

            # Otherwise, get URLs from this sitemap
            urls = root.findall('.//ns:url/ns:loc', namespace)
            return [url.text for url in urls]

        except Exception as e:
            print(f"Error fetching sitemap: {e}")
            return []

    def filter_posts_by_year(self, urls: List[str], year: int) -> List[str]:
        """Filter URLs to only include posts from a specific year."""
        filtered = []
        for url in urls:
            # Try to extract year from URL pattern or fetch the page
            if f"/{year}/" in url or f"-{year}-" in url:
                filtered.append(url)
        return filtered

    def get_post_urls_from_year_sitemap(self, year: int, limit: int = None) -> List[str]:
        """Get post URLs from a specific year's sitemap."""
        all_urls = self.fetch_sitemap()

        # If we got a sitemap index, fetch the year-specific one
        year_sitemap = None
        for url in all_urls:
            if str(year) in url and 'sitemap' in url:
                year_sitemap = url
                break

        if year_sitemap:
            print(f"Fetching year-specific sitemap: {year_sitemap}")
            try:
                response = self.session.get(year_sitemap, timeout=self.timeout)
                response.raise_for_status()

                root = ET.fromstring(response.content)
                namespace = {'ns': 'http://www.sitemaps.org/schemas/sitemap/0.9'}
                urls = root.findall('.//ns:url/ns:loc', namespace)
                post_urls = [url.text for url in urls]

                if limit:
                    post_urls = post_urls[:limit]

                print(f"Found {len(post_urls)} posts from {year}")
                return post_urls
            except Exception as e:
                print(f"Error fetching year sitemap: {e}")

        # Fallback: filter from all URLs
        filtered = self.filter_posts_by_year(all_urls, year)
        if limit:
            filtered = filtered[:limit]
        return filtered

    def load_urls_from_file(self, file_path: str, limit: int = None) -> List[str]:
        """Load post URLs from a text file (one URL per line)."""
        print(f"Loading URLs from {file_path}...")

        try:
            with open(file_path, 'r') as f:
                urls = [line.strip() for line in f if line.strip() and line.strip().startswith('http')]

            if limit:
                urls = urls[:limit]

            print(f"Loaded {len(urls)} URLs from file")
            return urls
        except Exception as e:
            print(f"Error loading URLs from file: {e}")
            return []

    def fetch_page_with_playwright(self, url: str) -> str:
        """Fetch a page using Playwright (real browser) to bypass bot protection."""
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            try:
                page = browser.new_page()

                # Add cookie if provided (for authenticated access)
                if self.cookie:
                    # Parse the domain from base_url
                    from urllib.parse import urlparse
                    parsed = urlparse(self.base_url)
                    domain = parsed.netloc

                    page.context.add_cookies([{
                        'name': 'substack.sid',
                        'value': self.cookie,
                        'domain': domain,
                        'path': '/'
                    }])

                page.goto(url, wait_until='networkidle', timeout=self.timeout * 1000)
                # Wait a bit for any dynamic content to load
                time.sleep(1)
                content = page.content()
                return content
            finally:
                browser.close()

    def fetch_page_with_requests(self, url: str) -> str:
        """Fetch a page using requests/cloudscraper (fallback method)."""
        max_retries = 3
        for attempt in range(max_retries):
            response = self.session.get(url, timeout=self.timeout)

            if response.status_code == 200:
                return response.text
            elif response.status_code == 403 and attempt < max_retries - 1:
                wait_time = (attempt + 1) * 2
                print(f"    Got 403, waiting {wait_time}s before retry {attempt + 2}/{max_retries}...")
                time.sleep(wait_time)
            else:
                response.raise_for_status()

        # If we get here, return the last response anyway
        return response.text

    def extract_links_from_post(self, post_url: str) -> Tuple[str, List[str]]:
        """Extract all links from a post and return the post title."""
        print(f"  Extracting links from {post_url}...")

        try:
            # Fetch the page content using the appropriate method
            if self.use_playwright:
                html_content = self.fetch_page_with_playwright(post_url)
            else:
                html_content = self.fetch_page_with_requests(post_url)

            soup = BeautifulSoup(html_content, 'html.parser')

            # Get post title
            title_tag = soup.find('h1') or soup.find('title')
            title = title_tag.get_text(strip=True) if title_tag else "Unknown Title"

            # Extract all links from the post content
            # Substack posts are typically in article tags or divs with specific classes
            content_area = soup.find('article') or soup.find('div', class_=re.compile('post|article|content'))

            if content_area:
                links = [a['href'] for a in content_area.find_all('a', href=True)]
            else:
                # Fallback to all links
                links = [a['href'] for a in soup.find_all('a', href=True)]

            # Filter out internal navigation links, social media, etc.
            external_links = []
            for link in links:
                # Skip anchors, mailto, tel, etc.
                if link.startswith('#') or link.startswith('mailto:') or link.startswith('tel:'):
                    continue

                # Skip Substack internal links (comments, share, etc.)
                if 'substack.com' in link and any(x in link for x in ['/subscribe', '/comments', '/share']):
                    continue

                # Make relative URLs absolute
                if link.startswith('/') or not link.startswith('http'):
                    link = urljoin(post_url, link)

                external_links.append(link)

            # Remove duplicates while preserving order
            seen = set()
            unique_links = []
            for link in external_links:
                if link not in seen:
                    seen.add(link)
                    unique_links.append(link)

            print(f"    Found {len(unique_links)} unique links")
            return title, unique_links

        except Exception as e:
            print(f"    Error extracting links: {e}")
            return "Error fetching post", []

    def check_link(self, link: str) -> Tuple[bool, str]:
        """
        Check if a link is broken and return the error type.
        Returns: (is_broken, error_type)
        """
        try:
            # Special case: local.theonion.com
            if 'local.theonion.com' in link:
                return True, "SSL Error (local.theonion.com)"

            # Try to fetch the link
            response = self.session.get(link, timeout=self.timeout, allow_redirects=True)

            # Check for 404
            if response.status_code == 404:
                return True, "HTTP 404"

            # Check for other error status codes
            if response.status_code >= 400:
                return True, f"HTTP {response.status_code}"

            # Check for soft 404s in the page title
            soup = BeautifulSoup(response.text, 'html.parser')
            title = soup.find('title')
            if title:
                title_text = title.get_text().lower()
                if any(phrase in title_text for phrase in ['404 error', 'page not found', 'not found', '404']):
                    return True, "Soft 404 (page title contains '404' or 'not found')"

            # Link is working
            return False, "OK"

        except requests.exceptions.SSLError as e:
            return True, f"SSL Error: {str(e)[:100]}"
        except requests.exceptions.Timeout:
            return True, "Timeout"
        except requests.exceptions.ConnectionError as e:
            # Check for DNS failures
            if 'Name or service not known' in str(e) or 'nodename nor servname provided' in str(e):
                return True, "DNS Failure"
            return True, f"Connection Error: {str(e)[:100]}"
        except Exception as e:
            return True, f"Unknown Error: {str(e)[:100]}"

    def check_post_links(self, post_url: str):
        """Check all links in a post and record broken ones."""
        title, links = self.extract_links_from_post(post_url)

        print(f"  Checking {len(links)} links...")
        broken_count = 0

        for link in links:
            is_broken, error_type = self.check_link(link)

            if is_broken:
                broken_count += 1
                self.results.append({
                    'post_title': title,
                    'post_url': post_url,
                    'broken_link': link,
                    'error_type': error_type
                })
                print(f"    ✗ BROKEN: {link[:80]}... ({error_type})")

            # Be polite: add a small delay between requests
            time.sleep(1)  # Increased to 1s to avoid rate limiting

        print(f"  Found {broken_count} broken links in this post\n")

    def generate_report(self, output_file: str = 'broken_links_report.csv'):
        """Generate a CSV report of broken links."""
        if not self.results:
            print("No broken links found!")
            return

        print(f"\nGenerating report: {output_file}")

        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['post_title', 'post_url', 'broken_link', 'error_type'])
            writer.writeheader()
            writer.writerows(self.results)

        print(f"Report generated with {len(self.results)} broken links")

        # Also generate Excel file with hyperlinks if openpyxl is available
        if OPENPYXL_AVAILABLE:
            excel_file = output_file.replace('.csv', '.xlsx')
            self.generate_excel_report(excel_file)

    def generate_excel_report(self, output_file: str = 'broken_links_report.xlsx'):
        """Generate an Excel report with clickable hyperlinks."""
        if not self.results:
            return

        print(f"Generating Excel report: {output_file}")

        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Broken Links"

        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Write headers
        headers = ['Post Title', 'Post URL', 'Broken Link', 'Error Type']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Write data with hyperlinks
        for row_idx, result in enumerate(self.results, 2):
            # Post Title
            ws.cell(row=row_idx, column=1, value=result['post_title'])

            # Post URL with hyperlink
            cell = ws.cell(row=row_idx, column=2, value=result['post_url'])
            cell.hyperlink = result['post_url']
            cell.font = Font(color="0563C1", underline="single")

            # Broken Link with hyperlink
            cell = ws.cell(row=row_idx, column=3, value=result['broken_link'])
            cell.hyperlink = result['broken_link']
            cell.font = Font(color="0563C1", underline="single")

            # Error Type
            ws.cell(row=row_idx, column=4, value=result['error_type'])

        # Adjust column widths
        ws.column_dimensions['A'].width = 50  # Post Title
        ws.column_dimensions['B'].width = 60  # Post URL
        ws.column_dimensions['C'].width = 60  # Broken Link
        ws.column_dimensions['D'].width = 30  # Error Type

        # Save the workbook
        wb.save(output_file)
        print(f"Excel report generated: {output_file} (with clickable hyperlinks)")

    def run(self, year: int = None, limit: int = None, skip: int = 0, output_file: str = 'broken_links_report.csv', url_file: str = None):
        """Main entry point to run the link checker.

        Args:
            year: Year to check (if using sitemap)
            limit: Maximum number of posts to check (None = all remaining)
            skip: Number of posts to skip from the beginning (for batch processing)
            output_file: Output CSV filename
            url_file: Path to file containing URLs (one per line). If provided, year is ignored.
        """
        print(f"Substack Broken Link Checker")
        print(f"{'=' * 50}")
        print(f"Base URL: {self.base_url}")

        # Get post URLs
        if url_file:
            print(f"Mode: File input")
            print(f"URL file: {url_file}")
            if skip > 0:
                print(f"Skipping first {skip} posts")
            print(f"Post limit: {limit if limit else 'all remaining'}")
            post_urls = self.load_urls_from_file(url_file, limit=None)  # Load all first

            # Apply skip and limit
            if skip > 0:
                post_urls = post_urls[skip:]
            if limit:
                post_urls = post_urls[:limit]

        elif year:
            print(f"Mode: Sitemap")
            print(f"Year: {year}")
            print(f"Post limit: {limit if limit else 'all'}")
            post_urls = self.get_post_urls_from_year_sitemap(year, limit)
        else:
            print("Error: Must provide either 'year' or 'url_file'")
            return

        print(f"{'=' * 50}\n")

        if not post_urls:
            print("No posts found!")
            return

        # Check each post
        for i, post_url in enumerate(post_urls, 1):
            print(f"[{i}/{len(post_urls)}] Processing: {post_url}")
            self.check_post_links(post_url)

            # Add delay between posts to avoid rate limiting
            if i < len(post_urls):
                time.sleep(5)  # Increased to 5s to avoid rate limiting

        # Generate report
        self.generate_report(output_file)


def main():
    """Main function - supports command-line batch selection."""

    # Using authenticated session cookie to bypass bot protection
    cookie = None
    checker = SubstackLinkChecker('https://yoursubstack.substack.com', use_playwright=False, cookie=cookie)

    # Define batches (batch_number: (skip, limit, output_file))
    batches = {
        1:  (0,   25, 'broken_links_batch01.csv'),
        2:  (25,  25, 'broken_links_batch02.csv'),
        3:  (50,  25, 'broken_links_batch03.csv'),
        4:  (75,  25, 'broken_links_batch04.csv'),
        5:  (100, 25, 'broken_links_batch05.csv'),
        6:  (125, 25, 'broken_links_batch06.csv'),
        7:  (150, 25, 'broken_links_batch07.csv'),
        8:  (175, 25, 'broken_links_batch08.csv'),
        9:  (200, 25, 'broken_links_batch09.csv'),
        10: (225, 25, 'broken_links_batch10.csv'),
        11: (250, 25, 'broken_links_batch11.csv'),
        12: (275, 25, 'broken_links_batch12.csv'),
        13: (300, None, 'broken_links_batch13.csv'),  # Remaining posts
    }

    # Parse command line arguments
    if len(sys.argv) > 1:
        arg = sys.argv[1].lower()

        if arg in ['--all', '-a']:
            # Run all batches sequentially
            print("Running ALL batches (1-13). This will take 2-3 hours.\n")
            for batch_num in range(1, 14):
                skip, limit, output_file = batches[batch_num]
                print(f"\n{'='*60}")
                print(f"Starting Batch {batch_num}/13")
                print(f"{'='*60}\n")
                checker.run(url_file='post_urls_2020.txt', skip=skip, limit=limit, output_file=output_file)
                print(f"\n✓ Batch {batch_num} complete! Report saved to {output_file}")
                if batch_num < 13:
                    print("Taking a 5-second break before next batch...")
                    time.sleep(5)
            print("\n" + "="*60)
            print("ALL BATCHES COMPLETE!")
            print("="*60)

        elif arg in ['--help', '-h']:
            print("Usage:")
            print("  python substack_link_checker.py --batch 1    # Run batch 1 (posts 0-24)")
            print("  python substack_link_checker.py -b 5         # Run batch 5 (posts 100-124)")
            print("  python substack_link_checker.py --all        # Run all batches (2-3 hours)")
            print("  python substack_link_checker.py --help       # Show this help")
            print("\nAvailable batches: 1-13 (25 posts each, except batch 13 gets remaining)")

        elif arg in ['--batch', '-b']:
            if len(sys.argv) < 3:
                print("Error: Please specify a batch number (1-13)")
                print("Example: python substack_link_checker.py --batch 1")
                sys.exit(1)

            try:
                batch_num = int(sys.argv[2])
                if batch_num not in batches:
                    print(f"Error: Batch {batch_num} not found. Valid batches: 1-13")
                    sys.exit(1)

                skip, limit, output_file = batches[batch_num]
                print(f"Running Batch {batch_num}/13")
                checker.run(url_file='post_urls_2020.txt', skip=skip, limit=limit, output_file=output_file)
                print(f"\n✓ Batch {batch_num} complete! Report saved to {output_file}")

            except ValueError:
                print(f"Error: Invalid batch number '{sys.argv[2]}'. Must be 1-13")
                sys.exit(1)
        else:
            print(f"Unknown argument: {arg}")
            print("Use --help to see available options")
            sys.exit(1)
    else:
        # No arguments provided - show helpful message
        print("Substack Broken Link Checker - Batch Mode")
        print("=" * 60)
        print("\nUsage:")
        print("  python substack_link_checker.py --batch 1    # Run batch 1")
        print("  python substack_link_checker.py --batch 2    # Run batch 2")
        print("  python substack_link_checker.py --all        # Run all batches")
        print("  python substack_link_checker.py --help       # Show help")
        print("\nEach batch processes 25 posts and takes ~12-15 minutes")
        print("Total batches: 13 (covering 300+ posts)")
        print("\nRecommended: Run one batch at a time to monitor progress")
        print("=" * 60)


if __name__ == '__main__':
    main()

